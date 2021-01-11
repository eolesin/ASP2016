# ------------------------- Function -----------------------------------------------

# filter a Tab delimited data file with length column headers
# capture measurements for non-zero values in a Database
# Narrow reporting to the length range specified

# ------------------------- Import Libraries ---------------------------------------

import os, sys
import shutil
import sqlite3
import pandas as pd
import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import BarChart, Series, Reference

# ------------------------- User interface ------------------------------------------

# prompt user for input file name
in_file_name = input('Enter XLSX Data File Name :\n')

# in_file_name = 'HeightAU_SJB_2018_Sample_Descriptions.xlsx'

# prompt user for min and max for length range (WARNING no checking of input values !)
v1 = input('Enter Minimum Length value (default = 121) :\n')

try:
    vMin = int(v1)
except:
    vMin = 121

v2 = input('Enter Maximum Length value (default = 296) :\n')

try:
    vMax = int(v2)
except:
    vMax = 296

# Prompt user for percentage for first screen - percentage of minimum total RFU
v3 = input('Enter percentage for first screen (lowest total RFU) default = 2.5 :\n')

try:
    first_screen = float(v3)
except:
    first_screen = 2.5

# Prompy user for percentage for second screen - percentage of minimum total RFU
v4 = input('Enter percentage for second screen (individual total RFU) default = 3.0 :\n')

try:
    second_screen = float(v4)
except:
    second_screen = 3.0

out_file_name =  "(Bin_Out_"+str(first_screen)+"_"+str(second_screen)+")_"+in_file_name

# -------------------------- Create Database ---------------------------------------------

# create a database and a connection to it
connection = sqlite3.connect("arisa.db")    # connect to database - create if file doesn't exist
cursor = connection.cursor()                # cursor holds data going in or coming out of the database

cursor.execute("DROP TABLE IF EXISTS length_header;")  # remove any existing data in the database
cursor.execute("DROP TABLE IF EXISTS sample_label;")
cursor.execute("DROP TABLE IF EXISTS sample_data;")
cursor.execute("DROP TABLE IF EXISTS bin_data;")
connection.commit()                              # this sends the queued instructions to the database

# if intent was to create a database for more than one Arisa data file then
#    would need to add a data_file_index field to length_header and sample label tables

# store full header data (line 1 of file) for ability to recreate raw data

# create Length_Header table
cursor.execute("CREATE TABLE length_header (id int,length float)")
connection.commit()

# store Label text and (potentially parse out) location/sample_date (no fields for parse yet)

# create Sample_Label table
cursor.execute("CREATE TABLE sample_label (sample_number INTEGER PRIMARY KEY,label TEXT,arisa_text TEXT)")
connection.commit()

# store raw length (nnn.d)- intensity (nnnnn.ddd) pairs (sparse data - drop zero's)

# create Sample_Data table
cursor.execute("CREATE TABLE sample_data (data_number INTEGER PRIMARY KEY,sample_number INTEGER,length FLOAT,intensity FLOAT)")
connection.commit()

# create Bin_Data table for new bins created
cursor.execute("CREATE TABLE bin_data (data_number INTEGER PRIMARY KEY, sample_number INTEGER,bin_number INTEGER,bin INTEGER,intensity FLOAT,bin_range TEXT)")
connection.commit()

# --------------------------- Import Data into Tables ---------------------------------------------
# Read Arisa Workbook

xl = pd.ExcelFile(in_file_name)

# create 2 dataframes
df_in1 = xl.parse(0,0,None,0)
df_in2 = xl.parse(1)

# import the length headers
z = 1
for len_val in df_in1.index:
    cursor.execute("INSERT INTO length_header (id,length) VALUES (?,?)", (z, len_val));
    z = z + 1
connection.commit();        # never forget this (if you want the changes to be saved)

length_rows = z - 1

# get both the header that came with the arisa file (df_in1) and the descriptive text (df_in2)
z = 1
for site_val in df_in1.columns:
    cursor.execute("INSERT INTO sample_label (sample_number,arisa_text) VALUES (?,?)", (z, site_val));
    z = z + 1
connection.commit();

# the number of sample sites
sites = z - 1

#  descriptive text (df_in2)
z = 0
for site_val in df_in2.columns:
    if (z > 0):
        cursor.execute("UPDATE sample_label SET label = ? WHERE sample_number = ?", (site_val, z));
    z = z + 1
connection.commit();

for row_index,row in df_in1.iterrows():
    z = 1
    for rfu in row:
        if pd.notna(rfu):
            cursor.execute("INSERT INTO sample_data (sample_number,length,intensity) VALUES (?,?,?)", (z, row_index, rfu));
        z = z + 1

connection.commit();

# done reading Arisa file and writing data to database - close the input file
# how to close ???

# ------------------------------------- Spreadsheet work --------------------------------------------

# Copy Excel workbook template

# copy and open a workbook using a template

# copy the template file to a temporary working file
shutil.copy('Autobin_Template.xlsx','wk_template.xlsx')

# open the working file
wb = load_workbook(filename = 'wk_template.xlsx')

# assign worksheet tabs to short sheet names
ws1 = wb['Drop_Raw']
ws2 = wb['Raw_Sums']

new_tab3_name = 'Trim_RFUs_Minus_' + str(first_screen) + '%'
ws3 = wb['Trim_RFUs_Minus_']
ws3.title = new_tab3_name

ws4 = wb['New_Sums']

new_tab5_name = 'Max_Bins_Minus_' + str(second_screen) + '%'
ws5 = wb['Max_Bins']
ws5.title = new_tab5_name

ws6 = wb['Final_Data']
ws7 = wb['As_Percentage']
ws8 = wb['Charts']


# assign especially long cell formats to a shorter variable name
sample_label_border = Border(left=Side(border_style='thin',color='FF000000'),
                                          right=Side(border_style='thin',color='FF000000'),
                                          top=Side(border_style='thin',color='FF000000'),
                                          bottom=Side(border_style='thin',color='FF000000'))

strip_border = Border(top=Side(border_style='thin',color='FF000000'),
                                          bottom=Side(border_style='thin',color='FF000000'))

# subroutine to fill in site header - parameters are : starting row,column,worksheet# for example ws1
def site_header(my_row,my_col_offset,my_ws):
    # List Sample Site names horizontally
    for (col_num,site_label) in cursor.execute('SELECT sample_number,label FROM sample_label'):
        j = col_num - 1 + my_col_offset
        # the sample label text
        my_ws.cell(row=my_row, column=j).value = site_label
        # set alignment of text inside cell and text wrapping
        my_ws.cell(row=my_row, column=j).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # set cell background color
        my_ws.cell(row=my_row, column=j).fill = PatternFill(fgColor="c6e2ff", fill_type="solid")
        # set cell borders
        my_ws.cell(row=my_row, column=j).border = sample_label_border


# ----------------- fill in Drop_Raw tab (ws1)

# color header cells
# "Length" cell A1
ws1.cell(row=1, column=1).fill = PatternFill(fgColor="ffebcd", fill_type = "solid")

# List Sample Site names horizontally
site_header(1,2,ws1)

# create empty dictionaries - key = length, value = row number
all_lengths = dict()                # for all rows of data
trim_lengths = dict()               # non-zero only for rows within desired (Trim) length range
just_trim_lengths = dict()          # just_trim starting at 1, outside Trim range = 0

# List lengths vertically
z = 0
for (row_num, length) in cursor.execute('SELECT id,length FROM length_header'):
    # row_num = id (starts at 1), length = length

    # length header value
    ws1.cell(row=row_num+1, column=1).value = length
    # length header formating
    ws1.cell(row=row_num+1, column=1).number_format = '##,##0.0'
    ws1.cell(row=row_num+1, column=1).fill = PatternFill(fgColor="ffebcd", fill_type="solid")
    ws1.cell(row=row_num+1, column=1).border = sample_label_border
    ws1.cell(row=row_num+1, column=1).alignment = Alignment(horizontal='center')

    # capture length and row numbers in dictionaries
    all_lengths[length] = row_num                           # key = length, value = row_num
    if (length >= vMin and length <= vMax):                 # if inside trim range
        trim_lengths[length] = row_num                      # key = length, value =  row_num
        z = z + 1                                           # track row number within trim range
        just_trim_lengths[length] = z                       # key = length, value =  row_num within trim range
    else:
        trim_lengths[length] = 0                            # if not in trim range set row_num to zero
        just_trim_lengths[length] = 0                       # if not in trim range set row_num to zero

# Fill in all non-zero data from Sample_Data table
# col_num = column number of data, fragment = length of sample, rfu = intensity
for (col_num,fragment,rfu) in cursor.execute('SELECT sample_number,length,intensity FROM sample_data'):
    j = all_lengths[fragment]    # all_lengths - key = length, value = row #

    # adjust row and column taking headers into account
    ws1.cell(row=j+1, column=col_num+1).value = rfu                   # intensity measurement
    ws1.cell(row=j+1, column=col_num+1).number_format = '##,##0.0'    # number formating

# done with Drop_Raw (ws1)

# -------------------- Raw_Sums tab (ws2)

# color header cells from template
# "Raw Sum by Sample -->" cells A1 and B1
ws2.cell(row=1, column=1).fill = PatternFill(fgColor="ffe4c4", fill_type="solid")
ws2.cell(row=1, column=2).fill = PatternFill(fgColor="ffe4c4", fill_type="solid")
# "Trim Sum by Sample -->" cells A2 and B2
ws2.cell(row=2, column=1).fill = PatternFill(fgColor="fff0f5", fill_type="solid")
ws2.cell(row=2, column=2).fill = PatternFill(fgColor="fff0f5", fill_type="solid")
# "Raw Sum By Length" cell A3
ws2.cell(row=3, column=1).fill = PatternFill(fgColor="eee8dc", fill_type="solid")
# "Length + 4" cell B3
ws2.cell(row=3, column=2).fill = PatternFill(fgColor="c4f0d5", fill_type="solid")

# List Sample Site names horizontally (now a subroutine)
site_header(3,3,ws2)

# List lengths vertically
for (row_num, length) in cursor.execute('SELECT id,length FROM length_header'):
    # row_num = id (starts at 1), length = length

    ws2.cell(row=row_num+3, column=2).value = length + 4
    ws2.cell(row=row_num+3, column=2).number_format = '##,##0.0'
    ws2.cell(row=row_num+3, column=2).fill = PatternFill(fgColor="c4f0d5", fill_type="solid")
    ws2.cell(row=row_num+3, column=2).border = sample_label_border
    ws2.cell(row=row_num+3, column=2).alignment = Alignment(horizontal='center')
    # add border for grid row
    for x in range(1, sites + 1, 1):
        ws2.cell(row=row_num + 3, column=x + 2).border = sample_label_border
    # if in trim range
    if (length >= vMin and length <= vMax):
        # add color for grid row if in trim range
        for x in range(1,sites+1,1):
            ws2.cell(row=row_num+3, column=x+2).fill = PatternFill(fgColor="fff0f5", fill_type="solid")

# create empty dictionaries to tally sums
raw_sums_r = dict()             # row sums  - key = length, value = row sum
raw_sums_c = dict()             # column sums - key = column, value = column sum
trim_tot_c = dict()             # column sums for trimmed range - key = length, value = row sum

# create an empty list
rfu_minus_in = list()           # a list of all data samples in the trim range

# Fill in all non-zero data from Sample_Data table
# col_num = column number of data, fragment = length of sample, rfu = intensity
for (col_num,fragment,rfu) in cursor.execute('SELECT sample_number,length,intensity FROM sample_data'):
    j = all_lengths[fragment]         # key = length, value = row #

    # fill data in the grid
    ws2.cell(row=j+3, column=col_num+2).value = rfu
    ws2.cell(row=j+3, column=col_num+2).number_format = '##,##0.0'

    # add data sample to sums
    raw_sums_r[j] = raw_sums_r.get(j, 0) + rfu                          # this sums all intensities in rows
    raw_sums_c[col_num] = raw_sums_c.get(col_num, 0) + rfu              # this sums all intensities in columns
    # if in trim range
    if (fragment >= vMin and fragment <= vMax):
        # add data sample to trim column sums
        trim_tot_c[col_num] = trim_tot_c.get(col_num, 0) + rfu    # this sums all intensities in trim range columns
        row_num = just_trim_lengths[fragment]                           # key = length, value = row number within trim range
        rfu_minus_in.append([row_num, col_num, rfu])                    # add to a list for use in ws3 (row, col, intensity)

# fill in the raw row sums
for key in raw_sums_r:
    z = key     # key is the row number
    ws2.cell(row=z+3, column=1).value = raw_sums_r[key]
    ws2.cell(row=z+3, column=1).number_format = '##,##0.0'
    ws2.cell(row=z+3, column=1).alignment = Alignment(horizontal='right',indent = 1)
    ws2.cell(row=z+3, column=1).fill = PatternFill(fgColor="eee8dc", fill_type="solid")
    ws2.cell(row=z+3, column=1).border = sample_label_border

# fill in the raw column sums
for key in raw_sums_c:
    z = key     # key is the column number
    ws2.cell(row=1, column=z+2).value = raw_sums_c[key]
    ws2.cell(row=1, column=z+2).number_format = '##,##0'
    ws2.cell(row=1, column=z+2).fill = PatternFill(fgColor="ffe4c4", fill_type="solid")
    ws2.cell(row=1, column=z+2).border = sample_label_border

# fill in the trim column sums
for key in trim_tot_c:
    z = key     # key is the column number
    ws2.cell(row=2, column=z+2).value = trim_tot_c[key]
    ws2.cell(row=2, column=z+2).number_format = '##,##0'
    ws2.cell(row=2, column=z+2).fill = PatternFill(fgColor="fff0f5", fill_type="solid")
    ws2.cell(row=2, column=z+2).border = sample_label_border


# -------------------- Trim_RFUs_Minus_X% tab (ws3)

# fix values for Minus - first_screen = X
ws3.cell(row=3, column=1).value = str(first_screen) + '%'
ws3.cell(row=4, column=3).value = 'Sum Minus ' +  str(first_screen) + '%'
ws3.cell(row=5, column=2).value = 'Minus ' + str(first_screen) + '%'

# color header cells
# "Length + 4" cell C5
ws3.cell(row=5, column=3).fill = PatternFill(fgColor="C4F0D5", fill_type = "solid")
# "Raw Sums" cell A5
ws3.cell(row=5, column=1).fill = PatternFill(fgColor="eee8dc", fill_type = "solid")
# "Minus X%" cell B5
ws3.cell(row=5, column=2).fill = PatternFill(fgColor="e6e6fa", fill_type = "solid")
# "Sum Minus X% ->" cells B4 and C4
ws3.cell(row=4, column=1).border = strip_border
ws3.cell(row=4, column=1).fill = PatternFill(fgColor="e6e6fa", fill_type="solid")
ws3.cell(row=4, column=2).border = strip_border
ws3.cell(row=4, column=2).fill = PatternFill(fgColor="e6e6fa", fill_type="solid")
ws3.cell(row=4, column=3).border = strip_border
ws3.cell(row=4, column=3).fill = PatternFill(fgColor="e6e6fa", fill_type="solid")
# "Trim Sum" cell A2
ws3.cell(row=2, column=1).fill = PatternFill(fgColor="eee8dc", fill_type="solid")
ws3.cell(row=2, column=1).border = sample_label_border

# create empty dictionaries for sums
minus_tot_c = dict()            # column sums for dropped at 2.5%
minus_tot_r = dict()            # row sums for dropped at 2.5%

# Trim%_Totals
tt_Max = 0
tt_Min = 1000000

# fill in the trim column sums using data gathered in ws2
for key in trim_tot_c:
    z = key
    ws3.cell(row=2, column=z + 3).value = trim_tot_c[key]
    ws3.cell(row=2, column=z + 3).number_format = '##,##0'
    ws3.cell(row=2, column=z + 3).fill = PatternFill(fgColor="eee8dc", fill_type="solid")
    ws3.cell(row=2, column=z + 3).border = sample_label_border
    # update Min and Max sums
    tt_Min = trim_tot_c[key] if tt_Min > trim_tot_c[key] else tt_Min
    tt_Max = trim_tot_c[key] if tt_Max < trim_tot_c[key] else tt_Max

# fill min and max column sums and 2.5% values
ws3.cell(row=2, column=2).value = tt_Min
ws3.cell(row=2, column=2).number_format = '##,##0'
ws3.cell(row=2, column=2).fill = PatternFill(fgColor="eee8dc", fill_type="solid")
ws3.cell(row=2, column=2).border = sample_label_border

ws3.cell(row=3, column=2).value = tt_Min * first_screen/100
ws3.cell(row=3, column=2).number_format = '##,##0.00'

ws3.cell(row=2, column=3).value = tt_Max
ws3.cell(row=2, column=3).number_format = '##,##0'
ws3.cell(row=2, column=3).fill = PatternFill(fgColor="eee8dc", fill_type="solid")
ws3.cell(row=2, column=3).border = sample_label_border

ws3.cell(row=3, column=3).value = tt_Max * first_screen/100
ws3.cell(row=3, column=3).number_format = '##,##0.00'

# save min 2.5% value for min_cut_off
min_cut_off = tt_Min * first_screen/100
cut_cnt = 0
# using sample data captured in trim range in ws2 fill both retained and first_screen%_minus values (lighter colors)
for row_num,col_num,rfu in rfu_minus_in:
    ws3.cell(row=row_num + 5, column=col_num + 3).value = rfu
    ws3.cell(row=row_num + 5, column=col_num + 3).number_format = '##,##0.0'
    if (rfu < min_cut_off):
        ws3.cell(row=row_num+5, column=col_num+3).font = Font(color = 'cdc0b0')  # use light colored font for dropped data
        cut_cnt = cut_cnt + 1                                                    # keep tally of dropped data
        minus_tot_c[col_num] = minus_tot_c.get(col_num, 0) + rfu                 # column sums for dropped at first_screen
        minus_tot_r[row_num] = minus_tot_r.get(row_num, 0) + rfu                 # row sums for dropped at first_screen

# fill in minus 2.5% columns
for key in trim_tot_c:
     z = key
     ws3.cell(row=4, column=z+3).value = trim_tot_c[key] - minus_tot_c.get(key,0)     # subtract dropped from col sums
     ws3.cell(row=4, column=z+3).number_format = '##,##0'
     ws3.cell(row=4, column=z+3).fill = PatternFill(fgColor="e6e6fa", fill_type="solid")
     ws3.cell(row=4, column=z+3).border = sample_label_border

# fill in the sample labels horizontally
site_header(5,4,ws3)

for (id,length) in cursor.execute('SELECT id,length FROM length_header'):
    if (length >= vMin and length <= vMax):                     # trim range only
        # fill in the length + 4
        z = just_trim_lengths[length]                            # dictionary takes in length - returns row #
        ws3.cell(row=z+5, column=3).value = length + 4           # now listing Length + 4
        ws3.cell(row=z+5, column=3).number_format = '##,##0.0'
        ws3.cell(row=z+5, column=3).fill = PatternFill(fgColor="C4F0D5", fill_type="solid")
        ws3.cell(row=z+5, column=3).border = sample_label_border
        # fill in raw sums rows with data gathered in ws2 - sums for all lengths
        j = all_lengths[length]                                 # dictionary takes in length - returns row #
        if(raw_sums_r.get(j,0)>0):                              # use "get" in case key not found (will return 0)
            ws3.cell(row=z+5, column=1).value = raw_sums_r[j]   # dictionary takes in row - returns row sum
            ws3.cell(row=z+5, column=1).fill = PatternFill(fgColor="eee8dc", fill_type="solid")  # only color cells with data
            ws3.cell(row=z+5, column=1).border = sample_label_border
            ws3.cell(row=z+5, column=2).fill = PatternFill(fgColor="e6e6fa", fill_type="solid")
            ws3.cell(row=z+5, column=2).border = sample_label_border
            if((raw_sums_r[j] -  minus_tot_r.get(z,0))>0):
                ws3.cell(row=z+5, column=2).value = raw_sums_r[j] -  minus_tot_r.get(z,0)   # subtract dropped from row total
                ws3.cell(row=z+5, column=2).number_format = '##,##0.0'
                ws3.cell(row=z+5, column=2).alignment = Alignment(horizontal='right', indent=1)
        ws3.cell(row=z+5, column=1).number_format = '##,##0.0'
        ws3.cell(row=z+5, column=1).alignment = Alignment(horizontal='right', indent=1)

# -------------------- New_Sums tab (ws4)

# add more cursors so multiple selects or inserts can happen without losing data in cursor
connection2 = sqlite3.connect("arisa.db")
cursor2 = connection2.cursor()
connection3 = sqlite3.connect("arisa.db")
cursor3 = connection3.cursor()

# create a couple empty dictionaries
new_binz  = dict()
max_bin   = dict()      # dictionary stores max value for each column of each new data bin

# color header cells
ws4.cell(row=1, column=1).value = 'Minus ' + str(first_screen) + '% RFU by Length Sum'

# fill in the sample labels horizontally
site_header(1,6,ws4)

p = 0
rx4 = 1

for bin in range(vMin+4,vMax+4,1):
    x1 = float(bin) - .4
    x2 = float(bin) + .5
    minus_sum = 0
    binned_sum = 0
    set = False
    # select all distinct lengths of data_samples falling within +/- limits of new bin
    for distinct_length in cursor.execute('SELECT distinct length FROM sample_data WHERE (length + 4 >= ?  AND length + 4 <= ?  and intensity > ?) ORDER BY length',(x1, x2, min_cut_off)):
        set = True
        sample_length = distinct_length[0]
        ws4.cell(row=p + 2, column=4).value = sample_length + 4
        ws4.cell(row=p + 2, column=4).number_format = '##,##0.0'
        ws4.cell(row=p + 2, column=4).alignment = Alignment(horizontal='center')
        ws4.cell(row=p + 2, column=3).value = str(x1) + " - " + str(x2)
        ws4.cell(row=p + 2, column=3).alignment = Alignment(horizontal='center')
        p = p + 1

        for (col_num,rfu) in cursor2.execute('SELECT sample_number,intensity FROM sample_data WHERE (length == ? and intensity > ?)',(sample_length, min_cut_off)):
            minus_sum = minus_sum + rfu                         # add the rfu for this sample site into minus_sum
            ws4.cell(row=p + 1, column=col_num + 5).value =  rfu
            ws4.cell(row=p + 1, column=col_num + 5).number_format = '##,##0.0'
            if (new_binz.get((bin,col_num),0) < rfu):         # if the intensity for this bin,column is less than rfu
                if (new_binz.get((bin, col_num), 0) > 1.0):   # new bin contains a value but is being overwritten by a larger one
                    ws4.cell(row=p + 1, column=col_num + 5).fill = PatternFill(fgColor="ffe4e1", fill_type="solid")  #highlight the cell overwritting
                new_binz[bin,col_num] = rfu                  # overwrite dictionary with new value (at bin, column)

    if(set):                                                 # if data found in bin +/- range
        ws4.cell(row=p + 2, column=5).value = bin            # fill in the bin #
        for col_num in range(1,sites+1,1):                   # step across all columns
            largest_rfu = new_binz.get((bin,col_num),0)      # get the largest rfu for that column
            if (largest_rfu > 0):
                ws4.cell(row=p + 2, column=col_num + 5).value = largest_rfu
                binned_sum = binned_sum + largest_rfu
                ws4.cell(row=p + 2, column=col_num + 5).number_format = '##,##0.0'

                max_bin[col_num] = max_bin.get(col_num,0) + 1
                for px in range(1,sites+6,1):                # format entire width with the new bin format
                    ws4.cell(row=p + 2, column=px).fill = PatternFill(fgColor="C4F0D5", fill_type="solid")
                    ws4.cell(row=p + 2, column=px).border = strip_border
                    bin_rg = str(x1) + " - " + str(x2)
                # save each non-zero column data value in the bin_data table
                cursor3.execute("INSERT INTO bin_data (sample_number,bin_number,bin,intensity,bin_range) VALUES (?,?,?,?,?)", (col_num,rx4,bin,largest_rfu,bin_rg));
        connection3.commit();
        ws4.cell(row=p+2, column=2).value = binned_sum
        ws4.cell(row=p+2, column=2).number_format = '##,##0.0'
        ws4.cell(row=p+2, column=1).value = minus_sum
        ws4.cell(row=p+2, column=1).number_format = '##,##0.0'
        p = p + 1
        rx4 = rx4 + 1

# -------------------- Max_Bins tab (ws5)

max_rfu_c   = dict()      # use to sum the column RFU on the Max_Bins tab
minus_2_rfu_c   = dict()  # use to sum the column RFU on the Max_Bins tab - then subtract cutoffs
max_rfu_r   = dict()      # use to sum the column RFU on the Max_Bins tab
max_bin_len = dict()      # key = bin, value = row #
no_bins     = dict()      # bins per column

# color header cells
ws5.cell(row=1, column=1).fill = PatternFill(fgColor="C4F0D5", fill_type="solid")
ws5.cell(row=1, column=2).fill = PatternFill(fgColor="C4F0D5", fill_type="solid")
ws5.cell(row=2, column=1).fill = PatternFill(fgColor="C4F0D5", fill_type="solid")
ws5.cell(row=2, column=2).fill = PatternFill(fgColor="C4F0D5", fill_type="solid")
ws5.cell(row=3, column=1).fill = PatternFill(fgColor="fafad2", fill_type="solid")
ws5.cell(row=3, column=2).fill = PatternFill(fgColor="fafad2", fill_type="solid")
ws5.cell(row=4, column=1).fill = PatternFill(fgColor="eedd82", fill_type="solid")
ws5.cell(row=4, column=2).fill = PatternFill(fgColor="eedd82", fill_type="solid")
ws5.cell(row=5, column=1).fill = PatternFill(fgColor="eedd82", fill_type="solid")
ws5.cell(row=5, column=2).fill = PatternFill(fgColor="eedd82", fill_type="solid")
ws5.cell(row=6, column=1).fill = PatternFill(fgColor="f0ffff", fill_type="solid")
ws5.cell(row=6, column=2).fill = PatternFill(fgColor="C4F0D5", fill_type="solid")

# rewrite %
ws5.cell(row=3, column=2).value = 'Minus ' + str(second_screen) + '%'

# fill in the sample labels horizontally
site_header(6,3,ws5)

# fill in the new bin and bin ranges from the bin_data table
rx5 = 1
for (bin,bin_range) in cursor.execute('SELECT distinct bin, bin_range FROM bin_data order by bin'):
    ws5.cell(row=rx5+6, column=2).value = bin
    ws5.cell(row=rx5+6, column=2).alignment = Alignment(horizontal='center')
    ws5.cell(row=rx5+6, column=2).fill = PatternFill(fgColor="C4F0D5", fill_type="solid")
    ws5.cell(row=rx5+6, column=2).border = sample_label_border

    ws5.cell(row=rx5+6, column=1).value = bin_range
    ws5.cell(row=rx5+6, column=1).alignment = Alignment(horizontal='center')
    ws5.cell(row=rx5+6, column=1).fill = PatternFill(fgColor="f0ffff", fill_type="solid")
    ws5.cell(row=rx5+6, column=1).border = sample_label_border

    max_bin_len[bin] = rx5       # keep track of row number of each bin value
    rx5 = rx5 + 1

# fill in the bin_data samples
for (col_num,bin,rfu) in cursor.execute('SELECT sample_number,bin,intensity FROM bin_data'):
    row_num = max_bin_len.get(bin)
    ws5.cell(row=row_num + 6, column=col_num + 2).value = rfu
    ws5.cell(row=row_num + 6, column=col_num + 2).number_format = '##,##0.0'
    max_rfu_c[col_num] = max_rfu_c.get(col_num,0) + rfu             # sum the RFU for each column
    minus_2_rfu_c[col_num] = minus_2_rfu_c.get(col_num,0) + rfu     # sum the RFU for each column
    max_rfu_r[row_num] = max_rfu_r.get(row_num, 0) + rfu            # sum the RFU for each row
    no_bins[col_num] = no_bins.get(col_num,0) + 1                   # count the number of bins with data in each column

# fill in the Total RFU and Y% data
for key in max_rfu_c:
    cx5 = key
    ws5.cell(row=2, column=cx5 + 2).value = max_rfu_c[key]
    ws5.cell(row=2, column=cx5 + 2).number_format = '##,##0.0'
    ws5.cell(row=2, column=cx5 + 2).border = sample_label_border
    ws5.cell(row=2, column=cx5 + 2).fill = PatternFill(fgColor="C4F0D5", fill_type="solid")
    ws5.cell(row=3, column=cx5 + 2).value = max_rfu_c[key] * second_screen/100
    ws5.cell(row=3, column=cx5 + 2).number_format = '##,##0.0'
    ws5.cell(row=3, column=cx5 + 2).border = sample_label_border
    ws5.cell(row=3, column=cx5 + 2).fill = PatternFill(fgColor="fafad2", fill_type="solid")

# fill in the number of bins in the column before 3% cut off
for key in no_bins:
    cx5 = key
    ws5.cell(row=1, column=cx5 + 2).value = no_bins[key]
    ws5.cell(row=1, column=cx5 + 2).border = sample_label_border
    ws5.cell(row=1, column=cx5 + 2).fill = PatternFill(fgColor="C4F0D5", fill_type="solid")

second_cut = 0
# set a lighter font for all data less than Y% value
for (col_num,bin,rfu) in cursor.execute('SELECT sample_number,bin,intensity FROM bin_data'):
    row_num = max_bin_len.get(bin)
    col_cut_off = max_rfu_c[col_num] * second_screen/100
    if (rfu < col_cut_off):
        ws5.cell(row=row_num + 6, column=col_num + 2).font = Font(color='cdc0b0')
        no_bins[col_num] = no_bins.get(col_num, 0) - 1                  # decrement the number of bins
        minus_2_rfu_c[col_num] = minus_2_rfu_c.get(col_num, 0) - rfu    # subtract the RFU for each column
        max_rfu_r[row_num] = max_rfu_r.get(row_num, 0) - rfu            # subtract the RFU for each row
        second_cut = second_cut + 1                                     # count the number of samples cut

# fill in the number of bins in the column after 3% cut off
for key in no_bins:
    cx5 = key
    ws5.cell(row=4, column=cx5 + 2).value = no_bins[key]
    ws5.cell(row=4, column=cx5 + 2).border = sample_label_border
    ws5.cell(row=4, column=cx5 + 2).fill = PatternFill(fgColor="eedd82", fill_type="solid")

# fill in the Total RFU after 3% cut
for key in minus_2_rfu_c:
    cx5 = key
    ws5.cell(row=5, column=cx5 + 2).value = minus_2_rfu_c[key]
    ws5.cell(row=5, column=cx5 + 2).number_format = '##,##0.0'
    ws5.cell(row=5, column=cx5 + 2).border = sample_label_border
    ws5.cell(row=5, column=cx5 + 2).fill = PatternFill(fgColor="eedd82", fill_type="solid")

# -------------------- Final_Data tab (ws6)  - add Minus Y%

# color header cells
ws6.cell(row=1, column=1).fill = PatternFill(fgColor="fafad2", fill_type="solid")
ws6.cell(row=1, column=2).fill = PatternFill(fgColor="eedd82", fill_type="solid")


# fill in the sample labels horizontally
site_header(1,3,ws6)

rx6 = 1
for (bin,bin_range) in cursor.execute('SELECT distinct bin, bin_range FROM bin_data order by bin'):
    row_num = max_bin_len.get(bin)
    if ((max_rfu_r.get(row_num, 0)) > 1.0):                   # data still in this bin
        ws6.cell(row=rx6 + 1, column=2).value = bin
        ws6.cell(row=rx6 + 1, column=2).fill = PatternFill(fgColor="eedd82", fill_type="solid")
        ws6.cell(row=rx6 + 1, column=2).alignment = Alignment(horizontal='center')
        ws6.cell(row=rx6 + 1, column=2).border = sample_label_border

        ws6.cell(row=rx6 + 1, column=1).value = bin_range
        ws6.cell(row=rx6 + 1, column=1).fill = PatternFill(fgColor="fafad2", fill_type="solid")
        ws6.cell(row=rx6 + 1, column=1).alignment = Alignment(horizontal='center')
        ws6.cell(row=rx6 + 1, column=1).border = sample_label_border


        for (col_num2, rfu) in cursor2.execute('SELECT sample_number,intensity FROM bin_data WHERE bin = ?', (bin,)):
                if(rfu > max_rfu_c[col_num2] * (second_screen/100)):
                    ws6.cell(row=rx6 + 1, column=col_num2 + 2).value = rfu
                    ws6.cell(row=rx6 + 1, column=col_num2 + 2).number_format = '##,##0.0'
        rx6 = rx6 + 1

# -------------------- 100%_Data tab (ws7)

# color header cells
ws7.cell(row=1, column=1).fill = PatternFill(fgColor="fafad2", fill_type="solid")
ws7.cell(row=1, column=2).fill = PatternFill(fgColor="eedd82", fill_type="solid")

# fill in the sample labels horizontally
site_header(1,3,ws7)

rx7 = 1
for (bin,bin_range) in cursor.execute('SELECT distinct bin, bin_range FROM bin_data order by bin'):
    row_num = max_bin_len.get(bin)
    if ((max_rfu_r.get(row_num, 0)) > 1.0):                   # data still in this bin
        ws7.cell(row=rx7 + 1, column=2).value = bin
        ws7.cell(row=rx7 + 1, column=2).fill = PatternFill(fgColor="eedd82", fill_type="solid")
        ws7.cell(row=rx7 + 1, column=2).alignment = Alignment(horizontal='center')
        ws7.cell(row=rx7 + 1, column=2).border = sample_label_border

        ws7.cell(row=rx7 + 1, column=1).value = bin_range
        ws7.cell(row=rx7 + 1, column=1).fill = PatternFill(fgColor="fafad2", fill_type="solid")
        ws7.cell(row=rx7 + 1, column=1).alignment = Alignment(horizontal='center')
        ws7.cell(row=rx7 + 1, column=1).border = sample_label_border


        for (col_num2, rfu) in cursor2.execute('SELECT sample_number,intensity FROM bin_data WHERE bin = ?', (bin,)):
                if(rfu > max_rfu_c[col_num2] * (second_screen/100)):
                    ws7.cell(row=rx7 + 1, column=col_num2 + 2).value = (100 * rfu)/minus_2_rfu_c[col_num2]
                    ws7.cell(row=rx7 + 1, column=col_num2 + 2).number_format = '##,##0.0'
        rx7 = rx7 + 1

# -------------------- Charts tab (ws8)


# move all data below row 34 - and transpose rows and columns
# fill in the sample labels horizontally

v_offset = 34

for (col_num, site_label) in cursor.execute('SELECT sample_number,label FROM sample_label'):
    j = col_num + v_offset
    # the sample label text
    ws8.cell(row=j, column=1).value = site_label

cx8 = 1
for (bin,bin_range) in cursor.execute('SELECT distinct bin, bin_range FROM bin_data order by bin'):
    row_num = max_bin_len.get(bin)
    if ((max_rfu_r.get(row_num, 0)) > 1.0):                   # data still in this bin
        ws8.cell(row=v_offset, column=1 + cx8).value = bin_range

        for (col_num2, rfu) in cursor2.execute('SELECT sample_number,intensity FROM bin_data WHERE bin = ?', (bin,)):
                if(rfu > max_rfu_c[col_num2] * (second_screen/100)):
                    ws8.cell(row=v_offset + col_num2, column=cx8 + 1).value = round((100 * rfu)/minus_2_rfu_c[col_num2],1)
        cx8 = cx8 + 1

# calculate placement of data - link to chart

ws8.column_dimensions['A'].width = 50
chart1 = BarChart()
chart1.type = "col"
chart1.style = 2
chart1.width = 0.7 * sites
chart1.height = 15
chart1.gapWidth = 10
chart1.overlap = 100
chart1.grouping = "percentStacked"
chart1.legend = None

# chart1.title = 'Percent Stacked Chart'
# chart1.y_axis.title = 'Percentage'
# chart1.x_axis.title = 'Sample Sites'
# chart1.varyColors = True
# chart1.shape = 4

site_list = Reference(ws8, min_col=1, min_row=35, max_row=35 + sites )
data = Reference(ws8, min_col=1, min_row=34, max_row=34 + sites, max_col=cx8)

chart1.add_data(data, titles_from_data=True)
chart1.set_categories(site_list)
#chart1.add_data(data)
ws8.add_chart(chart1, "A1")

# -------------------- Write the Excel output file - close down and report cutoff stats ------

# save the spreadsheet - close connections - print summary items
wb.save(out_file_name)

connection.close()
connection2.close()
connection3.close()

print(str(cut_cnt) + " data values found below " + str(first_screen) + "% cutoff")
print(str(second_cut) + " additional data values found below " + str(second_screen) + "% cutoff")
